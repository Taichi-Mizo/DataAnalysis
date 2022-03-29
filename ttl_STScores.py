'''
・各ワークシートごとに一行づつリストで取得して、Numpy Arrayに変換→pd.DataFrameに変換。
・「チーム名」カラムからユニークな値を取得。
・取得したチーム名をもとに、「チーム名」カラムが同じ値の行を取得。
・取得したDataFrameにおいて、「スコア」カラムの値が最大であるレコードを取得。
'''
from email import header
import openpyxl
import numpy as np
import pandas as pd

def createMaxDF(ws):
     #ws: worksheetオブジェクト
    rowsList = []
    for row in ws.iter_rows(min_row=1):
        valueList = []
        for num in range(6):
            valueList.append(row[num].value)
        rowsList.append(valueList)
    arrayList = np.array(rowsList)
    df = pd.DataFrame(arrayList)
    df.columns = df.iloc[0]
    #一行目をヘッダーとして指定したため、一行目のれこーどを削除する。
    df1 = df.drop(0)
    #「チーム名」カラムからユニークな値を取得。
    teamNames = df1['チーム名'].unique()

    maxRecList = []
    for tname in teamNames:
        df_team = df1[df1['チーム名']==tname]
        #取得したDataFrameにおいて、「スコア」カラムの値が最大であるレコードを取得
        max_score = float(df_team['スコア'].max())
        max_record = df_team[df_team['スコア'] == max_score]
        max_array = max_record.to_numpy()
        maxRecList.append(max_array[0])
    maxArrayList = np.array(maxRecList)
    #各チームの最大値をまとめたDataFrame
    maxDF = pd.DataFrame(maxArrayList)
    #ヘッダーは従来のものを使用
    maxDF.columns = df.iloc[0]
    
    return maxDF


def controller():
    wbpath = r"C:\Users\mttest1\Documents\PyScripts\20220328_STスコア.xlsx"
    wb = openpyxl.load_workbook(wbpath, data_only=True)
    #第0シートは、親シートのため第1シートからデータ取得
    wsObj = wb.worksheets[1:]

    for ws in wsObj:
        #シートごとの各チームの最大値を含むレコードを取得
        maxDF = createMaxDF(ws)
        #抽出したデータを既存のExcelに出力
        sheetNames = wb.sheetnames[1:]
        for sname in sheetNames:
            with pd.ExcelWriter(wbpath, mode='a', engine='openpyxl') as file:
                maxDF.to_excel(file, sheet_name=(f'max_{sname}'))
    wb.save(wbpath)
    print('処理が完了しました。')
    exit()
    
#実行
controller()

